import openpyxl


class ExcelReader:

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

    def read_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def close(self):
        self.workbook.close()


# Test code
# Example usage:
excel_file_path = "/Users/megha/PycharmProject/pywebautomationframework/src/utils/test_data.xlsx"
excel_1 = "/Users/megha/PycharmProject/pywebautomationframework/src/utils/test_megha.xlsx"

excel_reader = ExcelReader(excel_file_path)
excel_reader = ExcelReader(excel_1)
sheet_names = excel_reader.get_sheet_names()
sheet_1 = excel_reader.get_sheet_names()
print("Sheet names:", sheet_names)
print("sheet names:",sheet_1)

sheet_name = sheet_names[0]  # Choose the sheet you want to read
sheet_data = excel_reader.read_sheet(sheet_name)
print(f"Data from '{sheet_name}' sheet:")
for row in sheet_data:
    print(row)

excel_reader.close()
