import time
import pytest

from selenium import webdriver
from src.pageObjects.loginPage import LoginPage
from src.utils.ExcelRader import ExcelReader
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


def test_get_data():
    workbook = load_workbook("/Users/megha/PycharmProject/pywebautomationframework/src/utils/test_data.xlsx")
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    print("Data from Excel:", data)
    return data


class TestVWOLogin:

    @pytest.mark.usefixtures("setup")
    @pytest.mark.parametrize("username, password,result", test_get_data())
    #   def test_vwologin(self, setup):
    def test_vwo_login(self, username, password, result, setup):
        driver = setup
        loginPage = LoginPage(driver)
        loginPage.login_to_vwo(username, password)

        if result == "fail":
            assert "Your email, password, IP address or location did not match" in loginPage.get_errormessage().text
            driver.quit()
        else:
            assert "https://app.vwo.com/#/dashboard" in driver.current_url
